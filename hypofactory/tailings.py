"""Парсер отчётов института по хвостам обогатительных фабрик (xlsx).

Формат отчёта (см. «Как читать отчет института по хвостам.docx»):
- блок переработки: шихта руд, материал пруда-накопителя, отвальные хвосты;
- потоки хвостов: породные и (опционально) пирротиновые;
- по каждому потоку: гранулометрия по классам крупности (+125 ... -10 мкм)
  и минералогический разбор потерь внутри каждого класса
  (раскрытый/закрытый Pnt/Cp, примесь в пирротине, силикатная форма,
  пирит, миллерит, ...), плюс строки «извлекаемый / не извлекаемый металл».

«Элемент 28» = Ni, «Элемент 29» = Cu (анонимизация в исходных данных).
Парсер устойчив к #REF!, пустым ячейкам и вариациям написания классов.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

import openpyxl


# ------------------------------------------------------------------ model
@dataclass
class MineralRow:
    name: str
    ni_share_pct: float | None = None   # доля потерь Ni в классе, %
    ni_t: float | None = None           # Ni, т
    cu_share_pct: float | None = None
    cu_t: float | None = None


@dataclass
class SizeClass:
    label: str                          # например "-71 +45"
    share_pct: float | None = None      # доля класса в потоке, %
    ni_share_pct: float | None = None   # доля потерь Ni по классам, %
    ni_t: float | None = None
    cu_share_pct: float | None = None
    cu_t: float | None = None
    minerals: list[MineralRow] = field(default_factory=list)
    recoverable_ni_t: float | None = None
    recoverable_cu_t: float | None = None
    unrecoverable_ni_t: float | None = None
    unrecoverable_cu_t: float | None = None


@dataclass
class TailingsStream:
    name: str                           # "породные" | "пирротиновые"
    smt: float | None = None            # сухие метрические тонны
    ni_pct: float | None = None
    ni_t: float | None = None
    cu_pct: float | None = None
    cu_t: float | None = None
    classes: list[SizeClass] = field(default_factory=list)

    @property
    def total_recoverable_ni_t(self) -> float:
        return sum(c.recoverable_ni_t or 0 for c in self.classes)

    @property
    def total_recoverable_cu_t(self) -> float:
        return sum(c.recoverable_cu_t or 0 for c in self.classes)


@dataclass
class TailingsReport:
    source_file: str
    plant: str = ""                     # название фабрики (из имени файла)
    feed_smt: float | None = None       # переработано, СМТ
    feed_ni_pct: float | None = None
    feed_ni_t: float | None = None
    feed_cu_pct: float | None = None
    feed_cu_t: float | None = None
    tailings_smt: float | None = None   # отвальные хвосты всего
    tailings_ni_t: float | None = None
    tailings_cu_t: float | None = None
    streams: list[TailingsStream] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        for s, sd in zip(self.streams, d["streams"]):
            sd["total_recoverable_ni_t"] = s.total_recoverable_ni_t
            sd["total_recoverable_cu_t"] = s.total_recoverable_cu_t
        return d


# ------------------------------------------------------------------ helpers
_CLASS_RE = re.compile(r"^[+\-]\s*\d+")


def _num(v: Any) -> float | None:
    """Ячейка → число; #REF!, текст, пустота → None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _norm_class(label: str) -> str:
    """' -20 + 10' → '-20 +10'."""
    s = re.sub(r"\s+", " ", label.replace("мкм", "")).strip()
    s = re.sub(r"([+\-])\s+(\d)", r"\1\2", s)
    return s


def _is_class_label(s: str) -> bool:
    return bool(_CLASS_RE.match(s.strip()))


# ------------------------------------------------------------------ parser
def parse_tailings_xlsx(path: str | Path) -> TailingsReport:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Итог"] if "Итог" in wb.sheetnames else wb.worksheets[0]

    rows: list[list[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]
    report = TailingsReport(source_file=path.name, plant=_plant_from_name(path.name))

    # ---- шапка: переработка и отвальные хвосты
    for row in rows:
        cells = [_text(c) for c in row]
        joined = " ".join(cells).lower()
        nums = [_num(c) for c in row]
        vals = [n for n in nums if n is not None]
        if "итого" in joined and report.feed_smt is None and len(vals) >= 5:
            report.feed_smt, report.feed_ni_pct, report.feed_ni_t, \
                report.feed_cu_pct, report.feed_cu_t = vals[:5]
        if "отвальные хвосты" in joined and len(vals) >= 5 and report.tailings_smt is None:
            report.tailings_smt = vals[0]
            report.tailings_ni_t = vals[2]
            report.tailings_cu_t = vals[4]

    # ---- потоки: строки "Хвосты породные"/"Хвосты пирротиновые" с 5+ числами
    stream_starts: list[tuple[int, str]] = []
    for i, row in enumerate(rows):
        first_texts = [_text(c) for c in row[:4]]
        label = next((t for t in first_texts if t), "").lower()
        vals = [n for n in (_num(c) for c in row) if n is not None]
        if label.startswith("хвосты породные") and len(vals) >= 5:
            stream_starts.append((i, "породные"))
        elif label.startswith("хвосты пирротиновые") and len(vals) >= 5:
            stream_starts.append((i, "пирротиновые"))
        elif label.startswith("хвосты отвальные") and len(vals) >= 5:
            stream_starts.append((i, "отвальные (сводно)"))

    for idx, (start, name) in enumerate(stream_starts):
        end = stream_starts[idx + 1][0] if idx + 1 < len(stream_starts) else len(rows)
        stream = _parse_stream(rows, start, end, name, report)
        report.streams.append(stream)

    if not report.streams:
        report.warnings.append("Не найдено ни одного потока хвостов — проверьте формат файла")
    return report


def _parse_stream(rows: list[list[Any]], start: int, end: int, name: str,
                  report: TailingsReport) -> TailingsStream:
    vals = [n for n in (_num(c) for c in rows[start]) if n is not None]
    stream = TailingsStream(name=name)
    if len(vals) >= 5:
        stream.smt, stream.ni_pct, stream.ni_t, stream.cu_pct, stream.cu_t = vals[:5]

    i = start + 1
    # --- таблица гранулометрии: после строки "Класс крупности"
    while i < end:
        cells = [_text(c) for c in rows[i]]
        if any("класс крупности" in c.lower() for c in cells):
            i += 1
            break
        i += 1
    while i < end:
        cells = [_text(c) for c in rows[i]]
        label = next((c for c in cells if c), "")
        if "итого" in label.lower():
            i += 1
            break
        if _is_class_label(label):
            label_idx = next(j for j, c in enumerate(rows[i]) if _text(c))
            vals = [n for n in (_num(c) for c in rows[i][label_idx + 1:]) if n is not None]
            sc = SizeClass(label=_norm_class(label))
            if len(vals) >= 5:
                (sc.share_pct, sc.ni_share_pct, sc.ni_t,
                 sc.cu_share_pct, sc.cu_t) = vals[:5]
            stream.classes.append(sc)
        i += 1

    # --- минералогия по классам: заголовок "<класс> мкм"
    current: SizeClass | None = None
    while i < end:
        cells = [_text(c) for c in rows[i]]
        label = next((c for c in cells if c), "")
        low = label.lower()
        nums = [_num(c) for c in rows[i]]
        vals = [n for n in nums if n is not None]

        if _is_class_label(label):
            norm = _norm_class(label)
            current = next((c for c in stream.classes if c.label == norm), None)
            if current is None and "итого" not in low:
                # класс есть в минералогии, но отсутствовал в гранулометрии
                current = SizeClass(label=norm)
                stream.classes.append(current)
        elif current is not None and label and not low.startswith("итого"):
            if low.startswith("извлекаемый"):
                m = _row_metal_values(rows[i])
                current.recoverable_ni_t, current.recoverable_cu_t = m[1], m[3]
            elif low.startswith("не извлекаемый"):
                m = _row_metal_values(rows[i])
                current.unrecoverable_ni_t, current.unrecoverable_cu_t = m[1], m[3]
            elif not low.startswith(("потери", "свободный", "доля")):
                m = _row_metal_values(rows[i])
                if any(x is not None for x in m):
                    current.minerals.append(MineralRow(
                        name=label, ni_share_pct=m[0], ni_t=m[1],
                        cu_share_pct=m[2], cu_t=m[3]))
        elif low.startswith("итого извлекаемый") or low.startswith("итого не извлекаемый"):
            pass  # суммарные строки пересчитываем сами
        i += 1

    # проверка целостности
    for sc in stream.classes:
        got = sum(m.ni_t or 0 for m in sc.minerals)
        if sc.ni_t and got and abs(got - sc.ni_t) / max(sc.ni_t, 1e-9) > 0.05:
            report.warnings.append(
                f"{name}, класс {sc.label}: сумма минералогии Ni ({got:.0f} т) "
                f"расходится с итогом класса ({sc.ni_t:.0f} т)")
    return stream


def _row_metal_values(row: list[Any]) -> tuple[float | None, ...]:
    """Возвращает (ni_share, ni_t, cu_share, cu_t) из строки минералогии.

    В строке после текстовой метки идут до 4 числовых колонок; часть может
    отсутствовать (#REF!, пустые) — тогда позиции определяем по индексам
    непустых колонок в известной сетке (колонки D,E,F,G → 3..6 в 0-базе),
    с запасом ищем первые 4 числа после метки.
    """
    # индекс первой текстовой ячейки (метки)
    label_idx = next((j for j, c in enumerate(row) if _text(c)), 0)
    nums: list[float | None] = []
    for c in row[label_idx + 1:]:
        if _text(c) == "" and _num(c) is None:
            continue
        nums.append(_num(c))
    nums = (nums + [None] * 4)[:4]
    return tuple(nums)


def _plant_from_name(filename: str) -> str:
    base = re.sub(r"\.xlsx?$", "", filename, flags=re.I)
    base = re.sub(r"(?i)хвосты", "", base).strip(" _-")
    return base or filename


# ------------------------------------------------------------------ summary
def report_summary_text(r: TailingsReport) -> str:
    """Краткая текстовая сводка отчёта для промптов LLM."""
    lines = [f"Фабрика: {r.plant}. Отчёт: {r.source_file}."]
    if r.feed_smt:
        lines.append(f"Переработано {r.feed_smt:,.0f} СМТ (Ni {r.feed_ni_pct:.2f}%, "
                     f"Cu {r.feed_cu_pct:.2f}%).")
    if r.tailings_smt:
        lines.append(f"Отвальные хвосты: {r.tailings_smt:,.0f} СМТ, потери Ni "
                     f"{r.tailings_ni_t:,.0f} т, Cu {r.tailings_cu_t:,.0f} т.")
    for s in r.streams:
        lines.append(f"Поток «хвосты {s.name}»: {s.smt or 0:,.0f} СМТ, Ni {s.ni_t or 0:,.0f} т, "
                     f"Cu {s.cu_t or 0:,.0f} т; извлекаемый металл: Ni "
                     f"{s.total_recoverable_ni_t:,.0f} т, Cu {s.total_recoverable_cu_t:,.0f} т.")
        for c in s.classes:
            mins = ", ".join(f"{m.name}: Ni {m.ni_t or 0:,.0f} т / Cu {m.cu_t or 0:,.0f} т"
                             for m in c.minerals if (m.ni_t or 0) + (m.cu_t or 0) > 1)
            lines.append(f"  Класс {c.label} мкм ({c.share_pct or 0:.1f}% массы): "
                         f"Ni {c.ni_t or 0:,.0f} т, Cu {c.cu_t or 0:,.0f} т. {mins}")
    return "\n".join(lines)
